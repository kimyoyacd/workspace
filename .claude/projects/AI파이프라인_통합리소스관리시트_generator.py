# -*- coding: utf-8 -*-
"""AI 파이프라인 통합 리소스·스케줄 관리 시트 생성기.
동희실장 실별 MD 배분 + LINE 일자별 간트 + 밀리 유형별 표준소요일(마감 자동) 통합.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from datetime import date

FONT = "Arial"
# ---- palette
NAVY   = "1F2A44"
ACCENT = "FF6B35"
LBLUE  = "DCE6F1"
GREY   = "F2F2F2"
YEL    = "FFF2CC"   # 입력칸
GREEN  = "C6EFCE"
RED    = "FFC7CE"
YELLOWF= "FFEB9C"
GREYF  = "D9D9D9"
BARBL  = "9DC3E6"

thin = Side(style="thin", color="BFBFBF")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(sz=10, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c):
    return PatternFill("solid", fgColor=c)
def setcell(ws, coord, val, font=None, fillc=None, align=None, fmt=None, border=True):
    c = ws[coord]
    c.value = val
    c.font = font or F()
    if fillc: c.fill = fill(fillc)
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = box
    return c

CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENV= Alignment(horizontal="center", vertical="center")
LEFT= Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# =========================================================
# ④ 설정
# =========================================================
cf = wb.active
cf.title = "④설정"
cf.sheet_properties.tabColor = ACCENT
cf.column_dimensions["A"].width = 20
for col in "BCDEF": cf.column_dimensions[col].width = 14
cf.column_dimensions["D"].width = 18
cf.column_dimensions["G"].width = 16
for col in "HIJ": cf.column_dimensions[col].width = 14
cf.column_dimensions["L"].width = 12
cf.column_dimensions["N"].width = 22
cf.column_dimensions["O"].width = 10

setcell(cf, "A1", "기준일 (오늘·자동)", F(10,True), LBLUE)
setcell(cf, "B1", "=TODAY()", F(10,True), GREEN, CENV, "yyyy-mm-dd")
setcell(cf, "A2", "대상 월 (1일 날짜 입력)", F(10,True), LBLUE)
setcell(cf, "B2", date(2026,7,1), F(10,True,"0000FF"), YEL, CENV, "yyyy-mm-dd")
setcell(cf, "A3", "월 환산기준 (MD/MM)", F(10,True), LBLUE)
setcell(cf, "B3", 20, F(10,True,"0000FF"), YEL, CENV, "0")

# MD 환산표
setcell(cf, "A5", "■ 시간 → MD 환산  (근무 1일 = 1.0MD = 8시간)", F(10,True,"FFFFFF"), NAVY, LEFT)
cf.merge_cells("A5:B5")
setcell(cf, "A6", "시간", F(9,True), GREY, CENV)
setcell(cf, "B6", "MD", F(9,True), GREY, CENV)
for i in range(1,9):
    r = 6+i
    setcell(cf, f"A{r}", f"{i}h", F(9), None, CENV)
    setcell(cf, f"B{r}", round(i*0.125,3), F(9), None, CENV, "0.000")

# 유형별 표준 소요일
setcell(cf, "D5", "■ 업무유형별 표준 소요일 (영업일) — 마감 자동계산 근거", F(10,True,"FFFFFF"), NAVY, LEFT)
cf.merge_cells("D5:E5")
setcell(cf, "D6", "업무유형", F(9,True), GREY, CENV)
setcell(cf, "E6", "표준소요일", F(9,True), GREY, CENV)
types = [("시안(신규)",4),("사이즈베리",1),("다국어베리",1),("컬러베리",1),
         ("수정",1),("슬림이벤트",2),("베이직이벤트",4),("스페셜이벤트",6),
         ("영상제작(신규)",5),("영상베리(사이즈)",2)]
for i,(t,d) in enumerate(types):
    r = 7+i
    setcell(cf, f"D{r}", t, F(9), None, LEFT)
    setcell(cf, f"E{r}", d, F(9,True,"0000FF"), YEL, CENV, "0")
TYPE_RANGE = "$D$7:$D$16"; DAY_RANGE = "$E$7:$E$16"

# 실/팀 & 캐파
setcell(cf, "G5", "■ 실·팀별 월 캐파 (인원 × 1인 캐파)", F(10,True,"FFFFFF"), NAVY, LEFT)
cf.merge_cells("G5:J5")
for j,h in zip("GHIJ", ["실·팀","인원(TO)","1인월캐파(MD)","실 월캐파(MD)"]):
    setcell(cf, f"{j}6", h, F(9,True), GREY, CEN)
rooms = [("기획실",2,20),("디자인실",4,20),("MAX실(영상)",3,20),("VINA영상",2,20),("VINA이미지",2,20)]
for i,(nm,to,cap) in enumerate(rooms):
    r = 7+i
    setcell(cf, f"G{r}", nm, F(9), None, LEFT)
    setcell(cf, f"H{r}", to, F(9,True,"0000FF"), YEL, CENV, "0")
    setcell(cf, f"I{r}", cap, F(9,True,"0000FF"), YEL, CENV, "0")
    setcell(cf, f"J{r}", f"=H{r}*I{r}", F(9), None, CENV, "0")
ROOM_RANGE = "$G$7:$G$11"

# 상태값
setcell(cf, "L5", "■ 상태값", F(10,True,"FFFFFF"), NAVY, CEN)
states = ["대기","진행중","검수","보류","완료"]
for i,s in enumerate(states):
    setcell(cf, f"L{6+i}", s, F(9), None, CENV)
STATE_RANGE = "$L$6:$L$10"

# 비나 시나리오
setcell(cf, "N5", "■ 비나(VINA) 운영 시나리오", F(10,True,"FFFFFF"), NAVY, LEFT)
cf.merge_cells("N5:O5")
setcell(cf, "N6", "현재 적용 시나리오 (1~4)", F(9,True), LBLUE, LEFT)
setcell(cf, "O6", 1, F(11,True,"0000FF"), YEL, CENV, "0")
vscen = [("1","기획+내부디자인+비나영상 : 공수%로 기획/디자인/MAX 각각 분배"),
         ("2","비나=베리만 : 영상 사이즈베리는 디자인실이 진행"),
         ("3","비나 이미지팀 TO 추가발주 : 영상 전담 배치(추가비 발생)"),
         ("4","본사형 : 비나 영상팀+이미지팀 협업 (이미지 TO 추가 X)")]
for i,(n,d) in enumerate(vscen):
    setcell(cf, f"N{7+i}", n, F(9,True), None, CENV)
    setcell(cf, f"O{7+i}", d, F(9), None, LEFT)
cf.column_dimensions["O"].width = 52

setcell(cf, "A16", "※ 노란칸만 입력. 나머지는 자동계산.", F(8,False,"C00000"), None, LEFT, border=False)
cf.merge_cells("A16:B16")

# =========================================================
# ② 업무마스터 (통합 간트)
# =========================================================
wm = wb.create_sheet("②업무마스터")
wm.sheet_properties.tabColor = "1F2A44"
GDAYS = 31
FIRST_GANTT = 19  # S
headers = ["프로젝트","소재명 / 타이틀","유형","규격","주관실","상태","진척%",
           "시작일","소요일","마감예상","D-day","신호",
           "기획MD","디자인MD","MAX영상MD","VINA_MD","총MD","MM"]
widths  = [12,30,13,9,12,8,7, 11,7,11,8,8, 8,9,10,8,7,7]
for i,(h,w) in enumerate(zip(headers,widths)):
    col = get_column_letter(i+1)
    wm.column_dimensions[col].width = w

# Title row1
lastcol = get_column_letter(FIRST_GANTT+GDAYS-1)
setcell(wm, "A1", "AI 파이프라인 · 통합 리소스 & 스케줄 마스터   (한 행 = 한 업무 / 노란칸 입력 → 마감·D-day·간트 자동)",
        F(13,True,"FFFFFF"), NAVY, LEFT)
wm.merge_cells(f"A1:{lastcol}1")
wm.row_dimensions[1].height = 26

# group label row2
setcell(wm, "A2", "업무 정보", F(9,True,"FFFFFF"), ACCENT, CEN)
wm.merge_cells("A2:F2")
setcell(wm, "G2", "스케줄 (자동)", F(9,True,"FFFFFF"), "2E7D32", CEN)
wm.merge_cells("G2:L2")
setcell(wm, "M2", "실별 리소스 배분 (MD)", F(9,True,"FFFFFF"), "5B2C8D", CEN)
wm.merge_cells("M2:R2")
gs = get_column_letter(FIRST_GANTT); ge = lastcol
setcell(wm, f"{gs}2", "일자별 진행 (간트 · 자동)", F(9,True,"FFFFFF"), "0B6E99", CEN)
wm.merge_cells(f"{gs}2:{ge}2")

# header row3
for i,h in enumerate(headers):
    setcell(wm, f"{get_column_letter(i+1)}3", h, F(9,True,"FFFFFF"), NAVY, CEN)
# gantt date headers row3 (auto from config month)
for d in range(GDAYS):
    col = get_column_letter(FIRST_GANTT+d)
    f = (f'=IF( (\'④설정\'!$B$2+{d}) <= DATE(YEAR(\'④설정\'!$B$2),MONTH(\'④설정\'!$B$2)+1,0),'
         f' \'④설정\'!$B$2+{d}, "")')
    setcell(wm, f"{col}3", f, F(8,True,"FFFFFF"), NAVY, CENV, "d")
    wm.column_dimensions[col].width = 3.2
wm.row_dimensions[3].height = 22

# ---- data rows
FIRSTROW = 4
NROWS = 45
examples = [
    ("세나리버스","여름이벤트 KV 영상 제작","영상제작(신규)","30\"","MAX실(영상)","진행중",0.4,date(2026,7,2),2.5,0,0,3.0,0),
    ("세나리버스","여름이벤트 KV 사이즈베리","영상베리(사이즈)","6\"","VINA영상","대기",0.0,date(2026,7,3),0,0,0,0,1.5),
    ("칠대죄오리진","GOWTHER 라이브포스터","영상제작(신규)","30\"","MAX실(영상)","대기",0.0,date(2026,7,7),1.5,0.5,1.0,0,0),
    ("칠대죄오리진","DERIERI 엔드카드","시안(신규)","15\"","디자인실","진행중",0.6,date(2026,7,4),1.375,0.375,0.25,0,0),
    ("밀리의서재","7월 스페셜 이벤트","스페셜이벤트","-","기획실","검수",0.8,date(2026,7,1),1.0,3.0,0,0,0),
    ("밀리의서재","히어로배너 다국어베리","다국어베리","-","디자인실","완료",1.0,date(2026,6,30),0,0.5,0,0,0),
]
for idx in range(NROWS):
    r = FIRSTROW + idx
    ex = examples[idx] if idx < len(examples) else None
    # A~F inputs
    setcell(wm, f"A{r}", ex[0] if ex else None, F(9), None, LEFT)
    setcell(wm, f"B{r}", ex[1] if ex else None, F(9), None, LEFT)
    setcell(wm, f"C{r}", ex[2] if ex else None, F(9), None, CEN)
    setcell(wm, f"D{r}", ex[3] if ex else None, F(9), None, CENV)
    setcell(wm, f"E{r}", ex[4] if ex else None, F(9), None, CEN)
    setcell(wm, f"F{r}", ex[5] if ex else None, F(9), None, CENV)
    setcell(wm, f"G{r}", ex[6] if ex else None, F(9), None, CENV, "0%")
    setcell(wm, f"H{r}", ex[7] if ex else None, F(9,color="0000FF"), None, CENV, "yyyy-mm-dd")
    # I 소요일 (auto lookup, override 가능)
    setcell(wm, f"I{r}", f"=IFERROR(INDEX('④설정'!{DAY_RANGE},MATCH(C{r},'④설정'!{TYPE_RANGE},0)),\"\")",
            F(9), None, CENV, "0")
    # J 마감예상
    setcell(wm, f"J{r}", f"=IF(OR($H{r}=\"\",$I{r}=\"\"),\"\",WORKDAY($H{r},$I{r}-1))",
            F(9,True), None, CENV, "yyyy-mm-dd")
    # K D-day
    setcell(wm, f"K{r}", f"=IF($J{r}=\"\",\"\",IF($F{r}=\"완료\",\"완료\",$J{r}-'④설정'!$B$1))",
            F(9,True), None, CENV, "0")
    # L 신호
    setcell(wm, f"L{r}",
        (f"=IF($J{r}=\"\",\"\",IF($F{r}=\"완료\",\"✅완료\","
         f"IF($J{r}<'④설정'!$B$1,\"🔴지연\","
         f"IF($J{r}<='④설정'!$B$1+2,\"🔴임박\","
         f"IF($J{r}<='④설정'!$B$1+5,\"🟡주의\",\"🟢여유\")))))"),
        F(9,True), None, CENV)
    # M~P 실별 MD inputs
    for j,col in enumerate("MNOP"):
        v = ex[7+1+j] if ex else None  # ex indices: 8=기획,9=디자인,10=MAX,11=VINA
        setcell(wm, f"{col}{r}", v, F(9,color="0000FF"), None, CENV, "0.0##")
    # Q 총MD
    setcell(wm, f"Q{r}", f"=IF(SUM(M{r}:P{r})=0,\"\",SUM(M{r}:P{r}))", F(9,True), None, CENV, "0.0")
    # R MM
    setcell(wm, f"R{r}", f"=IF(Q{r}=\"\",\"\",Q{r}/'④설정'!$B$3)", F(9), None, CENV, "0.00")
    # gantt
    for d in range(GDAYS):
        col = get_column_letter(FIRST_GANTT+d)
        f = (f'=IF($H{r}="","",IF(AND({col}$3<>"",$H{r}<={col}$3,$J{r}>={col}$3),"■",""))')
        setcell(wm, f"{col}{r}", f, F(8,color=BARBL), None, CENV, border=False)

wm.freeze_panes = "C4"

# data validation
dv_type = DataValidation(type="list", formula1=f"'④설정'!{TYPE_RANGE}", allow_blank=True)
dv_room = DataValidation(type="list", formula1=f"'④설정'!{ROOM_RANGE}", allow_blank=True)
dv_stat = DataValidation(type="list", formula1=f"'④설정'!{STATE_RANGE}", allow_blank=True)
wm.add_data_validation(dv_type); wm.add_data_validation(dv_room); wm.add_data_validation(dv_stat)
dv_type.add(f"C{FIRSTROW}:C{FIRSTROW+NROWS-1}")
dv_room.add(f"E{FIRSTROW}:E{FIRSTROW+NROWS-1}")
dv_stat.add(f"F{FIRSTROW}:F{FIRSTROW+NROWS-1}")

# conditional formatting
grange = f"{get_column_letter(FIRST_GANTT)}{FIRSTROW}:{lastcol}{FIRSTROW+NROWS-1}"
wm.conditional_formatting.add(grange,
    FormulaRule(formula=[f'{get_column_letter(FIRST_GANTT)}{FIRSTROW}="■"'],
                fill=fill(BARBL), font=F(8,color=BARBL)))
lrange = f"L{FIRSTROW}:L{FIRSTROW+NROWS-1}"
wm.conditional_formatting.add(lrange, FormulaRule(formula=[f'ISNUMBER(SEARCH("지연",L{FIRSTROW}))'], fill=fill(RED)))
wm.conditional_formatting.add(lrange, FormulaRule(formula=[f'ISNUMBER(SEARCH("임박",L{FIRSTROW}))'], fill=fill(RED)))
wm.conditional_formatting.add(lrange, FormulaRule(formula=[f'ISNUMBER(SEARCH("주의",L{FIRSTROW}))'], fill=fill(YELLOWF)))
wm.conditional_formatting.add(lrange, FormulaRule(formula=[f'ISNUMBER(SEARCH("여유",L{FIRSTROW}))'], fill=fill(GREEN)))
wm.conditional_formatting.add(lrange, FormulaRule(formula=[f'ISNUMBER(SEARCH("완료",L{FIRSTROW}))'], fill=fill(GREYF)))
# 진척% bar-ish via 3color? skip

# =========================================================
# ③ 실별현황
# =========================================================
rs = wb.create_sheet("③실별현황")
rs.sheet_properties.tabColor = "5B2C8D"
for col,w in zip("ABCDEFGH",[16,14,12,10,12,10,10,10]): rs.column_dimensions[col].width = w
setcell(rs, "A1", "실·팀별 리소스 현황 (대상월 기준 · 자동집계)", F(13,True,"FFFFFF"), NAVY, LEFT)
rs.merge_cells("A1:H1"); rs.row_dimensions[1].height=24
setcell(rs, "A2", "대상 월", F(9,True), LBLUE)
setcell(rs, "B2", "='④설정'!B2", F(9,True), GREEN, CENV, "yyyy-mm")
hdr3 = ["실·팀","월 캐파(MD)","투입 MD","가동률","잔여 MD","투입 MM","진행중","지연"]
for i,h in enumerate(hdr3):
    setcell(rs, f"{get_column_letter(i+1)}3", h, F(9,True,"FFFFFF"), NAVY, CEN)
# rows: 기획실(M),디자인실(N),MAX실(O),VINA(P)
map_rows = [("기획실","M","='④설정'!J7"),
            ("디자인실","N","='④설정'!J8"),
            ("MAX실(영상)","O","='④설정'!J9"),
            ("VINA(영상+이미지)","P","='④설정'!J10+'④설정'!J11")]
ms = "'④설정'!$B$2"
me = "DATE(YEAR('④설정'!$B$2),MONTH('④설정'!$B$2)+1,0)"
for i,(nm,mdcol,cap) in enumerate(map_rows):
    r = 4+i
    setcell(rs, f"A{r}", nm, F(9,True), LBLUE, LEFT)
    setcell(rs, f"B{r}", cap, F(9), None, CENV, "0")
    setcell(rs, f"C{r}",
        f"=SUMIFS('②업무마스터'!${mdcol}$4:${mdcol}$48,'②업무마스터'!$J$4:$J$48,\">=\"&{ms},'②업무마스터'!$J$4:$J$48,\"<=\"&{me})",
        F(9), None, CENV, "0.0")
    setcell(rs, f"D{r}", f"=IFERROR(C{r}/B{r},0)", F(9,True), None, CENV, "0%")
    setcell(rs, f"E{r}", f"=B{r}-C{r}", F(9), None, CENV, "0.0")
    setcell(rs, f"F{r}", f"=C{r}/'④설정'!$B$3", F(9), None, CENV, "0.00")
    setcell(rs, f"G{r}", f"=COUNTIFS('②업무마스터'!$E$4:$E$48,A{r},'②업무마스터'!$F$4:$F$48,\"진행중\")", F(9), None, CENV, "0")
    setcell(rs, f"H{r}",
        f"=SUMPRODUCT(('②업무마스터'!$E$4:$E$48=A{r})*('②업무마스터'!$J$4:$J$48<>\"\")*('②업무마스터'!$J$4:$J$48<'④설정'!$B$1)*('②업무마스터'!$F$4:$F$48<>\"완료\"))",
        F(9), None, CENV, "0")
# 합계
tr = 4+len(map_rows)
setcell(rs, f"A{tr}", "합계", F(9,True,"FFFFFF"), ACCENT, CEN)
for col in "BCEF":
    setcell(rs, f"{col}{tr}", f"=SUM({col}4:{col}{tr-1})", F(9,True), "FDE9DF", CENV, "0.0" if col!="B" else "0")
setcell(rs, f"D{tr}", f"=IFERROR(C{tr}/B{tr},0)", F(9,True), "FDE9DF", CENV, "0%")
setcell(rs, f"G{tr}", f"=SUM(G4:G{tr-1})", F(9,True), "FDE9DF", CENV, "0")
setcell(rs, f"H{tr}", f"=SUM(H4:H{tr-1})", F(9,True), "FDE9DF", CENV, "0")
# 가동률 조건부서식
rs.conditional_formatting.add(f"D4:D{tr}", FormulaRule(formula=[f"D4>1"], fill=fill(RED)))
rs.conditional_formatting.add(f"D4:D{tr}", FormulaRule(formula=[f"AND(D4>0.8,D4<=1)"], fill=fill(YELLOWF)))
rs.conditional_formatting.add(f"D4:D{tr}", FormulaRule(formula=[f"AND(D4>0,D4<=0.8)"], fill=fill(GREEN)))
setcell(rs, f"A{tr+2}", "가동률 100%↑ 🔴과부하 · 80~100% 🟡타이트 · ~80% 🟢여유", F(8,False,"666666"), None, LEFT, border=False)
rs.merge_cells(f"A{tr+2}:H{tr+2}")

# =========================================================
# ① 대시보드
# =========================================================
db = wb.create_sheet("①대시보드")
db.sheet_properties.tabColor = ACCENT
wb.move_sheet("①대시보드", -(len(wb.sheetnames)-1))
for col,w in zip("ABCDEF",[22,14,14,14,14,14]): db.column_dimensions[col].width=w
setcell(db, "A1", "AI 파이프라인 리소스 대시보드", F(15,True,"FFFFFF"), NAVY, LEFT)
db.merge_cells("A1:F1"); db.row_dimensions[1].height=30
setcell(db, "A2", "기준일", F(9,True), LBLUE); setcell(db,"B2","='④설정'!B1",F(9),GREEN,CENV,"yyyy-mm-dd")
setcell(db, "C2", "대상월", F(9,True), LBLUE); setcell(db,"D2","='④설정'!B2",F(9),GREEN,CENV,"yyyy-mm")

R1="'②업무마스터'!$A$4:$A$48"; RJ="'②업무마스터'!$J$4:$J$48"; RF="'②업무마스터'!$F$4:$F$48"
RH="'②업무마스터'!$H$4:$H$48"; RQ="'②업무마스터'!$Q$4:$Q$48"
setcell(db, "A4", "■ 핵심 지표", F(11,True,"FFFFFF"), ACCENT, LEFT); db.merge_cells("A4:F4")
kpis = [
 ("등록 업무(건)", f"=COUNTA({R1})"),
 ("이번달 총 MD", f"=SUMIFS({RQ},{RJ},\">=\"&'④설정'!$B$2,{RJ},\"<=\"&DATE(YEAR('④설정'!$B$2),MONTH('④설정'!$B$2)+1,0))"),
 ("이번달 총 MM", "=B6/'④설정'!$B$3"),
 ("오늘 병렬진행", f"=SUMPRODUCT(({RH}<>\"\")*({RH}<='④설정'!$B$1)*({RJ}>='④설정'!$B$1)*({RF}<>\"완료\")*({R1}<>\"\"))"),
 ("🔴 지연", f"=SUMPRODUCT(({RJ}<>\"\")*({RJ}<'④설정'!$B$1)*({RF}<>\"완료\")*({R1}<>\"\"))"),
 ("🔴 임박(D-2)", f"=SUMPRODUCT(({RJ}>='④설정'!$B$1)*({RJ}<='④설정'!$B$1+2)*({RF}<>\"완료\")*({R1}<>\"\"))"),
 ("완료율", f"=IFERROR(COUNTIF({RF},\"완료\")/COUNTA({R1}),0)"),
]
rr = 5
for i,(lab,f) in enumerate(kpis):
    setcell(db, f"A{rr+i}", lab, F(10,True), LBLUE, LEFT)
    fmt = "0%" if lab=="완료율" else ("0.00" if "MM" in lab else "0.0" if "MD" in lab else "0")
    setcell(db, f"B{rr+i}", f, F(11,True,"C0392B" if "🔴" in lab else "000000"), None, CENV, fmt)
    db.merge_cells(f"B{rr+i}:C{rr+i}")

# 실별 가동률 요약 (링크)
setcell(db, "A14", "■ 실·팀 가동률 요약", F(11,True,"FFFFFF"), ACCENT, LEFT); db.merge_cells("A14:F14")
for i,h in enumerate(["실·팀","투입MD","캐파MD","가동률","신호"]):
    setcell(db, f"{get_column_letter(i+1)}15", h, F(9,True,"FFFFFF"), NAVY, CEN)
for i in range(4):
    r=16+i; sr=4+i
    setcell(db, f"A{r}", f"='③실별현황'!A{sr}", F(9,True), LBLUE, LEFT)
    setcell(db, f"B{r}", f"='③실별현황'!C{sr}", F(9), None, CENV, "0.0")
    setcell(db, f"C{r}", f"='③실별현황'!B{sr}", F(9), None, CENV, "0")
    setcell(db, f"D{r}", f"='③실별현황'!D{sr}", F(9,True), None, CENV, "0%")
    setcell(db, f"E{r}", f"=IF(D{r}>1,\"🔴과부하\",IF(D{r}>0.8,\"🟡타이트\",\"🟢여유\"))", F(9), None, CENV)
db.conditional_formatting.add("D16:D19", FormulaRule(formula=["D16>1"], fill=fill(RED)))
db.conditional_formatting.add("D16:D19", FormulaRule(formula=["AND(D16>0.8,D16<=1)"], fill=fill(YELLOWF)))
db.conditional_formatting.add("D16:D19", FormulaRule(formula=["AND(D16>0,D16<=0.8)"], fill=fill(GREEN)))
setcell(db, "A21", "※ 값 입력은 [②업무마스터] / [④설정] 에서. 이 화면은 전부 자동.", F(8,False,"666666"), None, LEFT, border=False)
db.merge_cells("A21:F21")

# =========================================================
# ⑤ 비나시뮬
# =========================================================
vs = wb.create_sheet("⑤비나시뮬")
vs.sheet_properties.tabColor = "0B6E99"
for col,w in zip("ABCDEFG",[26,16,16,16,16,16,16]): vs.column_dimensions[col].width=w
setcell(vs, "A1", "비나(VINA) 활용 시나리오 · 금액/공수 배분 시뮬레이터", F(13,True,"FFFFFF"), NAVY, LEFT)
vs.merge_cells("A1:G1"); vs.row_dimensions[1].height=24
# inputs
setcell(vs, "A3", "■ 입력 (노란칸)", F(10,True,"FFFFFF"), ACCENT, LEFT); vs.merge_cells("A3:C3")
inp = [("영상 프로젝트 총 견적(원)",92000000,"#,##0"),
       ("기획 공수 비중",0.2,"0%"),
       ("디자인 공수 비중",0.3,"0%"),
       ("영상 공수 비중",0.5,"0%"),
       ("비나 이미지 TO 추가발주비(원)",8000000,"#,##0")]
for i,(lab,v,fmt) in enumerate(inp):
    r=4+i
    setcell(vs, f"A{r}", lab, F(9,True), LBLUE, LEFT)
    setcell(vs, f"B{r}", v, F(9,True,"0000FF"), YEL, CENV, fmt)
setcell(vs, "A9", "합계 비중(=100% 확인)", F(9,True), None, LEFT)
setcell(vs, "B9", "=B5+B6+B7", F(9,True), None, CENV, "0%")

# scenario table
setcell(vs, "A11", "■ 시나리오별 실 배분 (금액)", F(10,True,"FFFFFF"), ACCENT, LEFT); vs.merge_cells("A11:G11")
cols = ["시나리오","기획실","디자인실","MAX실(영상)","VINA 정산","추가발주","합계검증"]
for i,h in enumerate(cols):
    setcell(vs, f"{get_column_letter(i+1)}12", h, F(9,True,"FFFFFF"), NAVY, CEN)
# amounts
# 1: 공수분배 — 기획=총*기획%, 디자인=총*디자인%, MAX=총*영상%, VINA정산=MAX실이 비나에 재하청(정보표기 0), 추가0
# 2: 비나=베리만 — 영상 신규는 MAX(총*영상%), 사이즈베리는 디자인실이 흡수(디자인=총*(디자인%+영상% *0.3)), 나머지 영상%*0.7 = MAX
# 3: 이미지TO 추가발주 — 1안 + 추가발주비 별도
# 4: 본사형 — 영상 전액 MAX 내 VINA 협업(추가발주 0), 기획/디자인 그대로
T="$B$4"; PK="$B$5"; PD="$B$6"; PV="$B$7"; ADD="$B$8"
rows5 = [
 ("① 공수 %분배",       f"={T}*{PK}", f"={T}*{PD}", f"={T}*{PV}", "0", "0"),
 ("② 비나=베리만",      f"={T}*{PK}", f"={T}*{PD}+{T}*{PV}*0.3", f"={T}*{PV}*0.7", "0", "0"),
 ("③ 이미지TO 추가발주", f"={T}*{PK}", f"={T}*{PD}", f"={T}*{PV}", "0", f"={ADD}"),
 ("④ 본사형(협업)",     f"={T}*{PK}", f"={T}*{PD}", f"={T}*{PV}", f"={T}*{PV}*0.4", "0"),
]
for i,(nm,k,d,mx,v,ad) in enumerate(rows5):
    r=13+i
    setcell(vs, f"A{r}", nm, F(9,True), LBLUE, LEFT)
    setcell(vs, f"B{r}", k, F(9), None, CENV, "#,##0")
    setcell(vs, f"C{r}", d, F(9), None, CENV, "#,##0")
    setcell(vs, f"D{r}", mx, F(9), None, CENV, "#,##0")
    setcell(vs, f"E{r}", v, F(9,color="777777"), None, CENV, "#,##0")
    setcell(vs, f"F{r}", ad, F(9), None, CENV, "#,##0")
    # 합계검증: 기획+디자인+MAX (+추가) vs 총액
    setcell(vs, f"G{r}", f"=B{r}+C{r}+D{r}+F{r}", F(9,True), None, CENV, "#,##0")
setcell(vs, "A18", "※ E열(VINA 정산)은 MAX실 금액 내부에서 비나에 지급되는 참고값(총액 합산 제외). ④본사형은 MAX실이 비나 협업으로 소화(추가발주 X).",
        F(8,False,"666666"), None, LEFT, border=False)
vs.merge_cells("A18:G18")
setcell(vs, "A19", "※ 현재 [④설정]에서 선택한 시나리오 번호에 맞는 행을 실 배분 기준으로 사용.",
        F(8,False,"666666"), None, LEFT, border=False)
vs.merge_cells("A19:G19")

# order tabs: 대시보드, 업무마스터, 실별현황, 설정, 비나시뮬
order = ["①대시보드","②업무마스터","③실별현황","④설정","⑤비나시뮬"]
wb._sheets.sort(key=lambda s: order.index(s.title))

out = "/tmp/claude-0/-home-user-workspace/95d017e2-5f62-5c9a-a2db-7dfa4dfc313a/scratchpad/AI파이프라인_통합리소스관리시트_v1.xlsx"
wb.save(out)
print("saved", out)
