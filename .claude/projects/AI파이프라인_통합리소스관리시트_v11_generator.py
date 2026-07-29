# -*- coding: utf-8 -*-
"""AI 파이프라인 통합 리소스/스케줄 관리시트 v11 생성기.

v10 대비 변경점
  1. 고객사 전달일(소재 단위 1개) 신설 + 전달일 여유·신호 자동
  2. 소재 = 3줄 고정 블록. 소재명/타이틀/전달일은 블록 첫줄만 입력, 아래 2줄 수식 자동 상속
  3. 수량(영상수/사이즈베리/다국어베리)은 블록 첫줄 전용 → 중복 집계 원천 차단
  4. 전달일 이후 날짜에 MD가 입력되면 해당 간트 셀 빨강 경고
  5. 데이터 300행(월 100소재)으로 확장

함수는 Google Sheets / Excel 공통 지원분만 사용(spilling 함수 미사용).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter as gl
from datetime import date

FONT = "맑은 고딕"
NAVY   = "1F2A44"
ACCENT = "FF6B35"
LBLUE  = "DCE6F1"
GREY   = "F2F2F2"
YEL    = "FFF2CC"   # 입력칸
INHERIT= "EFEFEF"   # 자동 상속(건드리지 말 것)
LOCK   = "E4E4E4"   # 블록 첫줄 전용칸의 나머지 줄
GREEN  = "C6EFCE"
RED    = "FFC7CE"
YELLOWF= "FFEB9C"
PURPLE = "D9C2E9"
BLUEF  = "BDD7EE"
DARKG  = "BFBFBF"
C_기획 = "FFF4E6"
C_이미지= "E8F4EC"
C_영상 = "E7EEF8"

thin = Side(style="thin", color="D0D0D0")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(sz=10, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c):
    return PatternFill("solid", fgColor=c)
CEN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENV = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFTV= Alignment(horizontal="left", vertical="center")

def setc(ws, coord, val, font=None, fillc=None, align=None, fmt=None, border=True):
    c = ws[coord]
    c.value = val
    c.font = font or F()
    if fillc: c.fill = fill(fillc)
    if align: c.alignment = align
    if fmt:   c.number_format = fmt
    if border: c.border = box
    return c

# ============================================================
# 레이아웃 상수
# ============================================================
TEAMS = ["마케팅1팀", "마케팅2팀", "마케팅3팀", "마케팅4팀", "기타"]
BLOCKS_PER_TEAM = 20          # 팀당 소재 20개
ROLES = ["기획", "이미지", "영상"]
ROWS_PER_BLOCK = 3
TEAM_ROWS = BLOCKS_PER_TEAM * ROWS_PER_BLOCK    # 60
FIRST_TEAM_HDR = 8                              # 첫 팀 헤더 행
GANTT_C0 = 21                                   # U열 = 1일
GDAYS = 31
LASTCOL = GANTT_C0 + GDAYS - 1                  # AY

# 팀 헤더/데이터 행 계산
team_layout = []   # (헤더행, 데이터시작, 데이터끝)
_r = FIRST_TEAM_HDR
for t in TEAMS:
    team_layout.append((t, _r, _r + 1, _r + TEAM_ROWS))
    _r = _r + 1 + TEAM_ROWS
DATA_FIRST = FIRST_TEAM_HDR + 1                 # 9
DATA_LAST  = team_layout[-1][3]                 # 312
HELPER_CAP = DATA_LAST + 3                      # 담당자 캐파 헬퍼
HELPER_HOL0 = DATA_LAST + 5                     # 공휴일 헬퍼 시작
HOLIDAYS = [
    ("신정", 1, 1), ("설날연휴", 2, 16), ("설날", 2, 17), ("설날연휴", 2, 18),
    ("삼일절(대체)", 3, 2), ("어린이날", 5, 5), ("부처님오신날(음력,확인필요)", 5, 24),
    ("현충일", 6, 6), ("광복절", 8, 15), ("추석연휴(음력,확인필요)", 9, 24),
    ("추석(음력,확인필요)", 9, 25), ("추석연휴(음력,확인필요)", 9, 26),
    ("개천절", 10, 3), ("한글날", 10, 9), ("성탄절", 12, 25),
]
TITLES = ["뱀피르", "RF 글로벌", "SOL 인챈트", "스톤에이지펫월드", "칠대죄오리진",
          "세나리버스", "몬길", "왕겜(GOT)", "몬스터길들이기", "라이즈오브"]
STATES = ["대기", "진행", "검수", "완료"]

# A~S 고정 영역
COLS = [
    ("A", "No",            5),
    ("B", "게임 타이틀",     14),
    ("C", "고객팀",          10),
    ("D", "소재명(업무명)",  28),
    ("E", "직군",             8),
    ("F", "담당자",          10),
    ("G", "고객사\n전달일",   11),
    ("H", "신호",             9),
    ("I", "전달일\n여유",      8),
    ("J", "소재\n총MD",       8),
    ("K", "소재\n최종완료",   11),
    ("L", "직군\nMD",         8),
    ("M", "직군\n시작일",     10),
    ("N", "직군\n완료일",     10),
    ("O", "영상\n수",          6),
    ("P", "사이즈\n베리",      6),
    ("Q", "다국어\n베리",      6),
    ("R", "상태",             8),
    ("S", "비고",            16),
]

wb = Workbook()

# ============================================================
# 사용법
# ============================================================
gd = wb.active
gd.title = "사용법"
gd.sheet_properties.tabColor = ACCENT
gd.column_dimensions["B"].width = 26
gd.column_dimensions["C"].width = 108
setc(gd, "B2", "AI 파이프라인 통합 리소스/스케줄 관리시트 v11", F(15, True, "FFFFFF"), NAVY, LEFTV, border=False)
gd.merge_cells("B2:C2"); gd.row_dimensions[2].height = 28
setc(gd, "B3", "한 소재 = 3줄 고정(기획/이미지/영상). 소재명·전달일은 첫줄만 입력하고 아래 2줄은 자동 상속됩니다.",
     F(10, False, "444444"), None, LEFTV, border=False)
gd.merge_cells("B3:C3")

sections = [
    ("1. 입력 규칙 - 어디에 뭘 넣나", [
        ("블록 첫줄(기획 줄)", "게임 타이틀 · 소재명 · 고객사 전달일 · 영상수/사이즈베리/다국어베리 · 상태. 이 6가지는 첫줄에만 넣습니다."),
        ("모든 줄", "담당자 + 오른쪽 날짜 칸의 MD. 직군마다 담당자가 다르므로 줄마다 각자 입력합니다."),
        ("자동 상속(회색칸)", "2·3줄의 게임타이틀/소재명/전달일은 첫줄을 그대로 따라옵니다. 수식이니 덮어쓰지 마세요."),
        ("직군 칸", "기획/이미지/영상이 미리 고정돼 있습니다. 안 쓰는 직군은 담당자·MD를 비워두면 집계에서 자동 제외됩니다."),
    ]),
    ("2. MD 입력 - 날짜 칸에 실제 값 그대로", [
        ("직접 입력", "그날 실제 투입 MD를 그 날짜 칸에 씁니다. 균등분산이 아닙니다. 예) 8/5 기획 0.5 / 8/5 이미지 0.4 / 8/6 영상 1.0 / 8/7 영상 1.0"),
        ("자동 계산", "직군MD·시작일·완료일은 그 줄에 넣은 날짜에서 자동으로 나옵니다. 소재 총MD와 최종완료일은 3줄을 묶어 첫줄에 표시됩니다."),
        ("시간→MD", "1h=0.13 / 2h=0.25 / 4h=0.5 / 8h=1.0 (근무 1일 = 1.0MD)"),
    ]),
    ("3. 고객사 전달일 - 이 시트의 기준선", [
        ("전달일 여유", "전달일 - 소재 최종완료일. 음수면 전달일을 넘긴 겁니다."),
        ("신호", "🔴지연(여유 음수) / 🟡임박(0~1일) / 🟢여유(2일 이상). 소재 블록 첫줄에 뜹니다."),
        ("빨간 날짜 칸", "전달일보다 뒤인 날짜에 MD를 넣으면 그 칸이 빨갛게 변합니다. 전달일을 넘겨 작업 중이라는 뜻입니다."),
    ]),
    ("4. 색으로 보는 신호", [
        ("파랑(상단)", "그 직군 정원(TO)의 80~100%가 찬 날"),
        ("보라(상단)", "정원 초과 - 배정 조율 필요"),
        ("빨강(날짜칸)", "같은 담당자가 같은 날 여러 소재 합쳐 1.0MD(8시간)를 넘김, 또는 전달일을 넘겨 작업 중"),
        ("회색(세로줄)", "주말 또는 대한민국 공휴일 - 워킹데이 아님"),
        ("줄 배경색", "기획=주황 / 이미지=초록 / 영상=파랑 계열로 직군을 구분합니다."),
    ]),
    ("5. 용량과 구조", [
        ("팀별 구간", "마케팅1~4팀 + 기타. 팀당 소재 20개(60행) 고정, 전체 월 100소재."),
        ("행이 모자라면", "그 팀 구간 안에서 3줄 단위로 삽입하고, 상단 합계와 연간요약의 참조 범위를 함께 넓혀야 합니다."),
        ("연간요약", "12개 월 탭에서 직접 집계합니다. 입력은 각 월 탭에서만 하세요."),
    ]),
]
r = 5
for title, items in sections:
    setc(gd, f"B{r}", title, F(11, True, "FFFFFF"), NAVY, LEFTV)
    gd.merge_cells(f"B{r}:C{r}")
    gd.row_dimensions[r].height = 20
    r += 1
    for k, v in items:
        setc(gd, f"B{r}", k, F(9, True), LBLUE, LEFTV)
        setc(gd, f"C{r}", v, F(9), None, LEFT)
        gd.row_dimensions[r].height = 26
        r += 1
    r += 1

# ============================================================
# 마스터
# ============================================================
ms = wb.create_sheet("마스터")
ms.sheet_properties.tabColor = "808080"
for col, w in zip("ABCDEFGHIJ", [22, 3, 14, 3, 10, 3, 10, 3, 26, 12]):
    ms.column_dimensions[col].width = w
setc(ms, "A1", "마스터 / 기준정보", F(14, True, "FFFFFF"), NAVY, LEFTV)
ms.merge_cells("A1:J1"); ms.row_dimensions[1].height = 24

setc(ms, "A2", "게임 타이틀", F(10, True), GREY, CENV)
setc(ms, "C2", "고객팀",     F(10, True), GREY, CENV)
setc(ms, "E2", "직군",       F(10, True), GREY, CENV)
setc(ms, "G2", "상태",       F(10, True), GREY, CENV)
setc(ms, "I2", "운영 파라미터", F(10, True), GREY, CENV)
setc(ms, "J2", "값",          F(10, True), GREY, CENV)
for i, t in enumerate(TITLES):
    setc(ms, f"A{3+i}", t, F(9, color="0000FF"), YEL, LEFTV)
for i, t in enumerate(TEAMS):
    setc(ms, f"C{3+i}", t, F(9), None, CENV)
for i, t in enumerate(ROLES):
    setc(ms, f"E{3+i}", t, F(9), None, CENV)
for i, t in enumerate(STATES):
    setc(ms, f"G{3+i}", t, F(9), None, CENV)

params = [("연도", 2026), ("월 가동일수", 20), ("담당자 1일 캐파(MD)", 1)]
for i, (k, v) in enumerate(params):
    setc(ms, f"I{3+i}", k, F(9, True), LBLUE, LEFTV)
    setc(ms, f"J{3+i}", v, F(9, True, "0000FF"), YEL, CENV, "0")

setc(ms, "I7", "직군 일 TO 정원", F(10, True), GREY, LEFTV)
setc(ms, "J7", "명", F(10, True), GREY, CENV)
for i, (k, v) in enumerate([("영상", 6), ("이미지", 8), ("기획", 3)]):
    setc(ms, f"I{8+i}", k, F(9, True), LBLUE, LEFTV)
    setc(ms, f"J{8+i}", v, F(9, True, "0000FF"), YEL, CENV, "0")

setc(ms, "I12", "대한민국 공휴일 (연도별 확인/수정 필요)", F(10, True), GREY, LEFTV)
setc(ms, "J12", "날짜", F(10, True), GREY, CENV)
for i, (nm, mm, dd) in enumerate(HOLIDAYS):
    setc(ms, f"I{13+i}", nm, F(9), None, LEFTV)
    setc(ms, f"J{13+i}", f"=DATE($J$3,{mm},{dd})", F(9, color="0000FF"), YEL, CENV, "yyyy-mm-dd")

setc(ms, "A30", "비나(VINA) 운영 시나리오 (참고)", F(11, True, "FFFFFF"), NAVY, LEFTV)
ms.merge_cells("A30:J30")
for i, (n, d) in enumerate([
    ("방안 1", "기획+내부디자인+비나영상 - 비나 영상공수는 영상실 매출 귀속"),
    ("방안 2", "비나=사이즈베리만 - 영상 베리는 디자인 흡수"),
    ("방안 3", "비나 이미지TO 추가발주+추가비용"),
    ("방안 4", "본사 프로세스 통합(이미지TO 추가X)")]):
    setc(ms, f"A{31+i}", n, F(9, True), LBLUE, LEFTV)
    setc(ms, f"C{31+i}", d, F(9), None, LEFTV)
    ms.merge_cells(f"C{31+i}:J{31+i}")

TITLE_DV  = "'마스터'!$A$3:$A$" + str(2 + len(TITLES))
STATE_DV  = "'마스터'!$G$3:$G$" + str(2 + len(STATES))

# ============================================================
# 월 탭 빌더
# ============================================================
SAMPLES = {
    8: [
        # (블록index, 게임타이틀, 소재명, 전달일(day), 수량(영상,사베,다베),
        #    {직군: (담당자, {day: md})})
        (0, "SOL 인챈트", "SOL_000120_카운트다운", 7, (1, 0, 0), {
            "기획":   ("최한은", {5: 0.5}),
            "이미지": ("김민성", {5: 0.4}),
            "영상":   ("강승일", {6: 1.0, 7: 1.0}),
        }),
        (1, "칠대죄오리진", "7DSOR_스토리트레일러", 7, (1, 2, 1), {
            "기획":   ("박태원", {5: 0.5}),
            "이미지": ("양재혁", {5: 1.0}),
            "영상":   ("이지현", {11: 1.5, 12: 1.0}),
        }),
        (2, "세나리버스", "TSKGB_엔드카드", 6, (1, 1, 0), {
            "영상":   ("강민우", {4: 1.0, 5: 1.0}),
        }),
    ]
}


def build_month(wb, m):
    ws = wb.create_sheet(f"{m}월")
    ws.sheet_properties.tabColor = NAVY

    # ---- 열 너비
    for col, _, w in COLS:
        ws.column_dimensions[col].width = w
    ws.column_dimensions["T"].width = 2
    for d in range(GDAYS):
        ws.column_dimensions[gl(GANTT_C0 + d)].width = 3.6

    LC = gl(LASTCOL)

    # ---- 1행 그룹 헤더
    setc(ws, "A1", f"{m}월 소재 정보  (소재 1개 = 기획/이미지/영상 3줄)", F(11, True, "FFFFFF"), NAVY, CEN)
    ws.merge_cells("A1:F1")
    setc(ws, "G1", "고객사 전달일 기준 (자동)", F(11, True, "FFFFFF"), "C0392B", CEN)
    ws.merge_cells("G1:K1")
    setc(ws, "L1", "직군별 집계 (자동)", F(11, True, "FFFFFF"), "2E7D32", CEN)
    ws.merge_cells("L1:N1")
    setc(ws, "O1", "수량 (소재당 1번)", F(11, True, "FFFFFF"), "5B2C8D", CEN)
    ws.merge_cells("O1:Q1")
    setc(ws, "R1", "관리", F(11, True, "FFFFFF"), "555555", CEN)
    ws.merge_cells("R1:S1")
    setc(ws, f"{gl(GANTT_C0)}1",
         f"{m}월 간트 - 날짜 칸에 실제 MD 직접 입력   |   상단 4줄 = 일자별 직군 TO 상태",
         F(11, True, "FFFFFF"), "0B6E99", CEN)
    ws.merge_cells(f"{gl(GANTT_C0)}1:{LC}1")
    ws.row_dimensions[1].height = 22

    # ---- 2행 컬럼 헤더 + 간트 날짜
    for col, name, _ in COLS:
        setc(ws, f"{col}2", name, F(9, True, "FFFFFF"), NAVY, CEN)
    for d in range(GDAYS):
        c = gl(GANTT_C0 + d)
        f = (f"=IF(MONTH(DATE('마스터'!$J$3,{m},1)+{d})={m},"
             f"DATE('마스터'!$J$3,{m},1)+{d},\"\")")
        setc(ws, f"{c}2", f, F(8, True, "FFFFFF"), NAVY, CENV, "d")
    ws.row_dimensions[2].height = 30

    # ---- 3~6행 일자별 TO 상태
    to_rows = [
        ("전체 TO (일자별 부하)", "=SUM('마스터'!$J$8:$J$10)", None),
        ("영상 TO",  "='마스터'!$J$8", "영상"),
        ("이미지 TO", "='마스터'!$J$9", "이미지"),
        ("기획 TO",  "='마스터'!$J$10", "기획"),
    ]
    for i, (label, quota, role) in enumerate(to_rows):
        r = 3 + i
        setc(ws, f"A{r}", label, F(9, True), LBLUE, LEFTV)
        ws.merge_cells(f"A{r}:F{r}")
        setc(ws, f"G{r}", "정원", F(9, True), GREY, CENV)
        setc(ws, f"H{r}", quota, F(9, True), GREY, CENV, "0.0")
        for c in "IJKLMNOPQRS":
            setc(ws, f"{c}{r}", None, F(9), GREY, CENV)
        for d in range(GDAYS):
            c = gl(GANTT_C0 + d)
            if role:
                f = (f'=SUMIF($E${DATA_FIRST}:$E${DATA_LAST},"{role}",'
                     f'{c}${DATA_FIRST}:{c}${DATA_LAST})')
            else:
                f = f"=SUM({c}${DATA_FIRST}:{c}${DATA_LAST})"
            setc(ws, f"{c}{r}", f, F(8, True), None, CENV, "0.0;;")
        ws.row_dimensions[r].height = 15
    setc(ws, "A7", "▼ 아래부터 입력. 노란칸=직접 입력 / 회색칸=자동(건드리지 마세요)",
         F(8, False, "C00000"), None, LEFTV, border=False)
    ws.merge_cells(f"A7:{LC}7")

    smp = {}
    for item in SAMPLES.get(m, []):
        smp[item[0]] = item

    # ---- 팀 구간 + 소재 블록
    gidx = 0
    for team, hdr_r, d0, d1 in team_layout:
        setc(ws, f"A{hdr_r}", f"{team}   ({d0}~{d1}행 · 소재 {BLOCKS_PER_TEAM}개)",
             F(10, True, "FFFFFF"), "4A5568", LEFTV)
        ws.merge_cells(f"A{hdr_r}:{LC}{hdr_r}")
        ws.row_dimensions[hdr_r].height = 18

        for b in range(BLOCKS_PER_TEAM):
            gidx += 1
            top = d0 + b * ROWS_PER_BLOCK
            sample = smp.get(gidx - 1) if team == TEAMS[0] else None
            _s_title = sample[1] if sample else None
            _s_name  = sample[2] if sample else None
            _s_due   = sample[3] if sample else None
            _s_qty   = sample[4] if sample else (None, None, None)
            _s_role  = sample[5] if sample else {}

            for k, role in enumerate(ROLES):
                r = top + k
                first = (k == 0)
                rc = {"기획": C_기획, "이미지": C_이미지, "영상": C_영상}[role]

                # A No
                setc(ws, f"A{r}", f'=IF($D${top}="","",{gidx})' if first else None,
                     F(9, color="808080"), rc, CENV, "0")
                # B 게임 타이틀
                if first:
                    setc(ws, f"B{r}", _s_title, F(9, color="0000FF"), YEL, LEFTV)
                else:
                    setc(ws, f"B{r}", f'=IF($B${top}="","",$B${top})',
                         F(9, color="909090"), INHERIT, LEFTV)
                # C 고객팀
                setc(ws, f"C{r}", team, F(9, color="808080"), rc, CENV)
                # D 소재명
                if first:
                    setc(ws, f"D{r}", _s_name, F(9, True, "0000FF"), YEL, LEFTV)
                else:
                    setc(ws, f"D{r}", f'=IF($D${top}="","",$D${top})',
                         F(9, color="909090"), INHERIT, LEFTV)
                # E 직군 (프리셋)
                setc(ws, f"E{r}", role, F(9, True), rc, CENV)
                # F 담당자
                nm = _s_role.get(role, (None, {}))[0]
                setc(ws, f"F{r}", nm, F(9, color="0000FF"), YEL, CENV)
                # G 고객사 전달일
                if first:
                    setc(ws, f"G{r}", date(2026, m, _s_due) if _s_due else None,
                         F(9, True, "0000FF"), YEL, CENV, "m/d")
                else:
                    setc(ws, f"G{r}", f'=IF($G${top}="","",$G${top})',
                         F(9, color="909090"), INHERIT, CENV, "m/d")
                # H 신호 / I 여유 / J 총MD / K 최종완료  (블록 첫줄만)
                if first:
                    setc(ws, f"H{r}",
                         f'=IF($I${top}="","",IF($I${top}<0,"🔴지연",'
                         f'IF($I${top}<=1,"🟡임박","🟢여유")))',
                         F(9, True), None, CENV)
                    setc(ws, f"I{r}",
                         f'=IF(OR($G${top}="",$K${top}=""),"",$G${top}-$K${top})',
                         F(9, True), None, CENV, "0;-0;0")
                    setc(ws, f"J{r}",
                         f'=IF(SUM($L${top}:$L${top+2})=0,"",SUM($L${top}:$L${top+2}))',
                         F(9, True), None, CENV, "0.0#")
                    setc(ws, f"K{r}",
                         f'=IF($J${top}="","",MAX($N${top}:$N${top+2}))',
                         F(9, True), None, CENV, "m/d")
                else:
                    for c in "HIJK":
                        setc(ws, f"{c}{r}", None, F(9), LOCK, CENV)
                # L 직군MD / M 시작 / N 완료  (줄마다)
                setc(ws, f"L{r}", f'=IF(SUM(U{r}:{LC}{r})=0,"",SUM(U{r}:{LC}{r}))',
                     F(9), rc, CENV, "0.0#")
                setc(ws, f"M{r}",
                     f'=IFERROR(INDEX(U$2:{LC}$2,MATCH(TRUE,INDEX(U{r}:{LC}{r}<>"",0),0)),"")',
                     F(9, color="606060"), rc, CENV, "m/d")
                setc(ws, f"N{r}",
                     f'=IFERROR(LOOKUP(2,1/(U{r}:{LC}{r}<>""),U$2:{LC}$2),"")',
                     F(9, color="606060"), rc, CENV, "m/d")
                # O/P/Q 수량 (블록 첫줄 전용)
                if first:
                    for c, v in zip("OPQ", _s_qty):
                        setc(ws, f"{c}{r}", v, F(9, color="0000FF"), YEL, CENV, "0;;")
                else:
                    for c in "OPQ":
                        setc(ws, f"{c}{r}", None, F(9), LOCK, CENV)
                # R 상태 / S 비고
                if first:
                    setc(ws, f"R{r}", "진행" if sample else None,
                         F(9, color="0000FF"), YEL, CENV)
                    setc(ws, f"S{r}", None, F(9), None, LEFTV)
                else:
                    setc(ws, f"R{r}", None, F(9), LOCK, CENV)
                    setc(ws, f"S{r}", None, F(9), None, LEFTV)

                # 간트 셀
                days = _s_role.get(role, (None, {}))[1]
                for d in range(GDAYS):
                    c = gl(GANTT_C0 + d)
                    setc(ws, f"{c}{r}", days.get(d + 1) if days else None,
                         F(8, color="0000FF"), None, CENV, "0.0#;;")
                ws.row_dimensions[r].height = 15

    # ---- 헬퍼 (조건부서식이 같은 시트 안에서만 참조하도록)
    setc(ws, f"A{HELPER_CAP}", "담당자 1일 캐파(MD)", F(8, color="909090"), None, LEFTV, border=False)
    setc(ws, f"B{HELPER_CAP}", "='마스터'!$J$5", F(8, color="909090"), None, CENV, "0.0", border=False)
    setc(ws, f"A{HELPER_HOL0 - 1}", "공휴일(자동 연동)", F(8, color="909090"), None, LEFTV, border=False)
    for i in range(len(HOLIDAYS)):
        setc(ws, f"A{HELPER_HOL0+i}", f"='마스터'!$I${13+i}", F(8, color="909090"), None, LEFTV, border=False)
        setc(ws, f"B{HELPER_HOL0+i}", f"='마스터'!$J${13+i}", F(8, color="909090"), None, CENV, "yyyy-mm-dd", border=False)
    HOL = f"$B${HELPER_HOL0}:$B${HELPER_HOL0+len(HOLIDAYS)-1}"
    CAP = f"$B${HELPER_CAP}"

    # ---- 드롭다운
    dv_t = DataValidation(type="list", formula1=TITLE_DV, allow_blank=True)
    dv_s = DataValidation(type="list", formula1=STATE_DV, allow_blank=True)
    ws.add_data_validation(dv_t); ws.add_data_validation(dv_s)
    dv_t.add(f"B{DATA_FIRST}:B{DATA_LAST}")
    dv_s.add(f"R{DATA_FIRST}:R{DATA_LAST}")

    # ---- 조건부서식 (추가 순서 = 우선순위)
    G0 = gl(GANTT_C0)
    top_rng  = f"{G0}3:{LC}6"
    day_rng  = f"{G0}3:{LC}{DATA_LAST}"
    cell_rng = f"{G0}{DATA_FIRST}:{LC}{DATA_LAST}"

    # 상단 TO 상태
    ws.conditional_formatting.add(top_rng, FormulaRule(
        formula=[f"AND({G0}3<>\"\",{G0}3>$H3)"], fill=fill(PURPLE), font=F(8, True, "4B0082")))
    ws.conditional_formatting.add(top_rng, FormulaRule(
        formula=[f"AND({G0}3>0,{G0}3>=$H3*0.8)"], fill=fill(BLUEF), font=F(8, True, "1F4E79")))

    # 달 범위 밖 열 / 주말·공휴일 회색  (간트 전체)
    ws.conditional_formatting.add(day_rng, FormulaRule(
        formula=[f'{G0}$2=""'], fill=fill(DARKG)))
    ws.conditional_formatting.add(day_rng, FormulaRule(
        formula=[f'AND({G0}$2<>"",OR(WEEKDAY({G0}$2,2)>5,COUNTIF({HOL},{G0}$2)>0))'],
        fill=fill(GREY)))

    # 전달일 초과 작업 (최우선 경고)
    ws.conditional_formatting.add(cell_rng, FormulaRule(
        formula=[f'AND({G0}{DATA_FIRST}<>"",$G{DATA_FIRST}<>"",{G0}$2<>"",{G0}$2>$G{DATA_FIRST})'],
        fill=fill(RED), font=F(8, True, "9C0006")))
    # 담당자 1일 캐파 초과
    ws.conditional_formatting.add(cell_rng, FormulaRule(
        formula=[f'AND($F{DATA_FIRST}<>"",{G0}{DATA_FIRST}<>"",'
                 f'SUMIF($F${DATA_FIRST}:$F${DATA_LAST},$F{DATA_FIRST},'
                 f'{G0}${DATA_FIRST}:{G0}${DATA_LAST})>{CAP})'],
        fill=fill("F8CBAD"), font=F(8, True, "C00000")))
    # 직군별 입력값 색
    for role, col in (("기획", C_기획), ("이미지", C_이미지), ("영상", C_영상)):
        ws.conditional_formatting.add(cell_rng, FormulaRule(
            formula=[f'AND($E{DATA_FIRST}="{role}",{G0}{DATA_FIRST}<>"")'],
            fill=fill(col), font=F(8, True, "1F2A44")))

    # 신호/여유 색
    for rng in (f"H{DATA_FIRST}:I{DATA_LAST}",):
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND($I{DATA_FIRST}<>"",$I{DATA_FIRST}<0)'], fill=fill(RED), font=F(9, True, "9C0006")))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND($I{DATA_FIRST}<>"",$I{DATA_FIRST}<=1)'], fill=fill(YELLOWF), font=F(9, True, "9C6500")))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'$I{DATA_FIRST}<>""'], fill=fill(GREEN), font=F(9, True, "006100")))

    ws.freeze_panes = "I9"
    return ws


for m in range(1, 13):
    build_month(wb, m)

# ============================================================
# 연간요약
# ============================================================
ys = wb.create_sheet("연간요약")
ys.sheet_properties.tabColor = ACCENT
for col, w in zip("ABCDEFGHIJKL", [3, 9, 11, 11, 11, 11, 12, 10, 10, 10, 10, 10]):
    ys.column_dimensions[col].width = w
setc(ys, "B2", "연간요약 - 월별 / 누적 / 가동률", F(14, True, "FFFFFF"), NAVY, LEFTV)
ys.merge_cells("B2:L2"); ys.row_dimensions[2].height = 24
setc(ys, "B3", "각 월 탭에서 직접 집계합니다. 입력은 해당 월 탭에서만 하세요.",
     F(9, False, "666666"), None, LEFTV, border=False)
ys.merge_cells("B3:L3")

hdr = ["월", "영상MD", "이미지MD", "기획MD", "합계MD", "누적MD",
       "영상수", "사이즈베리", "다국어베리", "소재수", "가동률"]
for i, h in enumerate(hdr):
    setc(ys, f"{gl(2+i)}5", h, F(9, True, "FFFFFF"), NAVY, CEN)

DF, DL = DATA_FIRST, DATA_LAST
for i in range(12):
    m = i + 1
    r = 6 + i
    t = f"'{m}월'"
    setc(ys, f"B{r}", f"{m}월", F(9, True), LBLUE, CENV)
    for j, role in enumerate(["영상", "이미지", "기획"]):
        setc(ys, f"{gl(3+j)}{r}",
             f'=SUMIF({t}!$E${DF}:$E${DL},"{role}",{t}!$L${DF}:$L${DL})',
             F(9), None, CENV, "0.0")
    setc(ys, f"F{r}", f"=SUM(C{r}:E{r})", F(9, True), None, CENV, "0.0")
    setc(ys, f"G{r}", f"=SUM($F$6:F{r})", F(9), None, CENV, "0.0")
    for j, col in enumerate("OPQ"):
        setc(ys, f"{gl(8+j)}{r}", f"=SUM({t}!${col}${DF}:${col}${DL})", F(9), None, CENV, "0;;")
    setc(ys, f"K{r}", f'=COUNTIF({t}!$A${DF}:$A${DL},">0")', F(9), None, CENV, "0;;")
    setc(ys, f"L{r}",
         f"=IFERROR(F{r}/(SUM('마스터'!$J$8:$J$10)*'마스터'!$J$4),0)",
         F(9, True), None, CENV, "0.0%")

tr = 18
setc(ys, f"B{tr}", "연간 합계", F(10, True, "FFFFFF"), ACCENT, CENV)
for c in "CDEF":
    setc(ys, f"{c}{tr}", f"=SUM({c}6:{c}17)", F(10, True), "FDE9DF", CENV, "0.0")
setc(ys, f"G{tr}", f"=F{tr}", F(10, True), "FDE9DF", CENV, "0.0")
for c in "HIJK":
    setc(ys, f"{c}{tr}", f"=SUM({c}6:{c}17)", F(10, True), "FDE9DF", CENV, "0;;")
setc(ys, f"L{tr}", f"=IFERROR(F{tr}/(SUM('마스터'!$J$8:$J$10)*'마스터'!$J$4*12),0)",
     F(10, True), "FDE9DF", CENV, "0.0%")

for lbl, rows, rr in (("상반기(1~6월)", "6:11", 20), ("하반기(7~12월)", "12:17", 21)):
    a, b_ = rows.split(":")
    setc(ys, f"B{rr}", lbl, F(9, True), LBLUE, CENV)
    for c in "CDEF":
        setc(ys, f"{c}{rr}", f"=SUM({c}{a}:{c}{b_})", F(9), None, CENV, "0.0")
    for c in "HIJK":
        setc(ys, f"{c}{rr}", f"=SUM({c}{a}:{c}{b_})", F(9), None, CENV, "0;;")

ys.conditional_formatting.add(f"L6:L{tr}", FormulaRule(formula=["L6>1"], fill=fill(RED)))
ys.conditional_formatting.add(f"L6:L{tr}", FormulaRule(formula=["AND(L6>0.8,L6<=1)"], fill=fill(YELLOWF)))
ys.conditional_formatting.add(f"L6:L{tr}", FormulaRule(formula=["AND(L6>0,L6<=0.8)"], fill=fill(GREEN)))

# 탭 순서
order = ["사용법", "마스터", "연간요약"] + [f"{m}월" for m in range(1, 13)]
wb._sheets.sort(key=lambda s: order.index(s.title))

out = "/home/user/workspace/.claude/projects/AI파이프라인_통합리소스관리시트_v11.xlsx"
wb.save(out)
print("saved:", out)
print(f"데이터행 {DATA_FIRST}~{DATA_LAST} / 팀 {len(TEAMS)} × 소재 {BLOCKS_PER_TEAM} = "
      f"{len(TEAMS)*BLOCKS_PER_TEAM}소재")
